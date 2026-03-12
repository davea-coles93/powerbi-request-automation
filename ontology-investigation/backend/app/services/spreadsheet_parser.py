"""Spreadsheet Parser — deterministic extraction of ontology elements from Excel/CSV.

Pass 1 of the two-pass ingestion architecture. Reads Excel (.xlsx/.xls) or CSV
files and extracts structured ontology elements using:
  - Frictionless Framework for robust type inference
  - Pandas for statistical profiling (cardinality, nulls, sample values)

The parser uses heuristics to classify columns as dimensions vs measures:
  - Numeric columns with high cardinality → likely measures
  - String/categorical columns with low cardinality → likely dimensions
  - Date/datetime columns → likely time dimensions

All extracted elements are tagged state="as-is" since they represent the
current state of the data source.

AI enrichment (entity grouping, relationship inference, metric creation)
happens in Pass 2 via the AI service — this module is purely deterministic.
"""

import csv
import hashlib
import io
import logging
import re
import tempfile
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)


# ── Helpers ──────────────────────────────────────────────────────────────


def _slug(name: str) -> str:
    """Generate a clean snake_case ID from a name."""
    slug = name.lower().strip()
    slug = re.sub(r"[^a-z0-9_\s-]", "", slug)
    slug = re.sub(r"[\s-]+", "_", slug)
    slug = re.sub(r"_+", "_", slug)
    return slug.strip("_")


def _display_name(raw: str) -> str:
    """Clean a column header into a display name."""
    name = raw.strip()
    # If already Title Case or mixed, keep it
    if name != name.lower() and name != name.upper():
        return name
    # Convert snake_case or UPPER to Title Case
    return name.replace("_", " ").strip().title()


def _infer_type_from_values(values: list) -> str:
    """Infer a data type from a sample of non-null values."""
    if not values:
        return "string"

    import datetime

    type_counts = {"number": 0, "datetime": 0, "boolean": 0, "string": 0}

    for v in values:
        if isinstance(v, bool):
            type_counts["boolean"] += 1
        elif isinstance(v, (int, float)):
            type_counts["number"] += 1
        elif isinstance(v, datetime.datetime):
            type_counts["datetime"] += 1
        elif isinstance(v, datetime.date):
            type_counts["datetime"] += 1
        else:
            sv = str(v).strip().lower()
            if sv in ("true", "false", "yes", "no", "1", "0"):
                type_counts["boolean"] += 1
            else:
                try:
                    float(sv.replace(",", ""))
                    type_counts["number"] += 1
                except (ValueError, AttributeError):
                    type_counts["string"] += 1

    # Return the dominant type
    return max(type_counts, key=type_counts.get)


def _classify_column(
    col_name: str,
    data_type: str,
    unique_count: int,
    total_count: int,
    null_count: int,
) -> str:
    """Classify a column as 'dimension', 'measure', or 'date_dimension'.

    Heuristics:
      - datetime → date_dimension
      - boolean → dimension
      - number + high cardinality (>50% unique) → measure
      - number + low cardinality (<20 unique) → dimension (e.g., Year, Month Number)
      - string + any cardinality → dimension
    """
    if data_type == "datetime":
        return "date_dimension"

    if data_type == "boolean":
        return "dimension"

    if data_type == "number":
        # Low cardinality numbers are likely dimension keys (Year, Month, etc.)
        if unique_count <= 20:
            return "dimension"
        # Medium cardinality — check ratio
        if total_count > 0 and unique_count / total_count < 0.3:
            return "dimension"
        return "measure"

    # String/other → always dimension
    return "dimension"


# ── Result Container ─────────────────────────────────────────────────────


class SpreadsheetParseResult:
    """Container for all extracted ontology elements from a spreadsheet."""

    def __init__(self):
        self.source_name: str = ""
        self.source_type: str = "spreadsheet"  # "excel" or "csv"
        self.sheets: list[dict] = []  # Per-sheet metadata
        self.entities: list[dict] = []
        self.attributes: list[dict] = []
        self.measures: list[dict] = []
        self.systems: list[dict] = []
        self.column_profiles: list[dict] = []  # Detailed profiling for review
        self.warnings: list[str] = []

    def to_dict(self) -> dict:
        return {
            "source_type": self.source_type,
            "source_name": self.source_name,
            "sheets": self.sheets,
            "entities": self.entities,
            "attributes": self.attributes,
            "measures": self.measures,
            "systems": self.systems,
            "column_profiles": self.column_profiles,
            "warnings": self.warnings,
        }


# ── Parser ───────────────────────────────────────────────────────────────


MAX_ROWS_FOR_PROFILING = 50_000  # Limit rows read for profiling (memory guard)
MAX_SAMPLE_VALUES = 10  # Number of sample values to include per column


class SpreadsheetParser:
    """Parse Excel or CSV files into ontology elements.

    Usage:
        parser = SpreadsheetParser()
        result = parser.parse_excel_bytes(data, filename="Sales.xlsx")
        # or
        result = parser.parse_csv_bytes(data, filename="Sales.csv")
    """

    def parse_excel_bytes(
        self,
        data: bytes,
        filename: str = "upload.xlsx",
        sheet_names: Optional[list[str]] = None,
        max_rows: int = MAX_ROWS_FOR_PROFILING,
    ) -> SpreadsheetParseResult:
        """Parse an Excel file from in-memory bytes.

        Args:
            data: Raw file bytes
            filename: Original filename (used for naming)
            sheet_names: Specific sheets to parse (None = all)
            max_rows: Maximum rows to read per sheet for profiling
        """
        result = SpreadsheetParseResult()
        result.source_name = Path(filename).stem
        result.source_type = "excel"

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(data),
                read_only=True,
                data_only=True,  # Resolve formulas to values
            )
        except Exception as e:
            result.warnings.append(f"Failed to open Excel file: {e}")
            return result

        try:
            sheets_to_parse = sheet_names or wb.sheetnames

            for sheet_name in sheets_to_parse:
                if sheet_name not in wb.sheetnames:
                    result.warnings.append(f"Sheet '{sheet_name}' not found")
                    continue

                ws = wb[sheet_name]
                self._parse_sheet(ws, sheet_name, result, max_rows)
        finally:
            wb.close()

        # Add the spreadsheet as a system
        self._add_system(result, filename)

        return result

    def parse_csv_bytes(
        self,
        data: bytes,
        filename: str = "upload.csv",
        encoding: str = "utf-8",
        delimiter: Optional[str] = None,
        max_rows: int = MAX_ROWS_FOR_PROFILING,
    ) -> SpreadsheetParseResult:
        """Parse a CSV file from in-memory bytes.

        Args:
            data: Raw file bytes
            filename: Original filename (used for naming)
            encoding: File encoding (default utf-8)
            delimiter: CSV delimiter (auto-detect if None)
            max_rows: Maximum rows to read for profiling
        """
        result = SpreadsheetParseResult()
        result.source_name = Path(filename).stem
        result.source_type = "csv"

        try:
            text = data.decode(encoding)
        except UnicodeDecodeError:
            # Fall back to latin-1 which never fails
            text = data.decode("latin-1")
            result.warnings.append(
                f"File not valid {encoding}, decoded as latin-1"
            )

        # Auto-detect delimiter
        if delimiter is None:
            try:
                sample = text[:8192]
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ","

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)

        try:
            headers = next(reader)
        except StopIteration:
            result.warnings.append("CSV file is empty")
            return result

        # Clean headers
        headers = [h.strip() for h in headers]
        if not any(headers):
            result.warnings.append("No column headers found")
            return result

        # Read rows into columns
        columns: dict[int, list] = {i: [] for i in range(len(headers))}
        row_count = 0

        for row in reader:
            if row_count >= max_rows:
                result.warnings.append(
                    f"File truncated at {max_rows} rows for profiling"
                )
                break
            for i, val in enumerate(row):
                if i < len(headers):
                    # Try to coerce types
                    columns[i].append(self._coerce_csv_value(val))
            row_count += 1

        # Build profiling data and ontology elements
        sheet_name = Path(filename).stem
        self._profile_columns(
            headers, columns, row_count, sheet_name, result,
        )

        result.sheets.append({
            "name": sheet_name,
            "row_count": row_count,
            "column_count": len(headers),
            "headers": headers,
        })

        # Add the CSV as a system
        self._add_system(result, filename)

        return result

    # ── Internal Methods ─────────────────────────────────────────────

    def _parse_sheet(
        self,
        ws,
        sheet_name: str,
        result: SpreadsheetParseResult,
        max_rows: int,
    ) -> None:
        """Parse a single Excel worksheet."""
        # Read headers from first row
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            result.warnings.append(f"Sheet '{sheet_name}' is empty")
            return

        headers = [str(h).strip() if h is not None else "" for h in header_row]

        # Skip sheets with no headers or only empty columns
        if not any(headers):
            result.warnings.append(f"Sheet '{sheet_name}' has no column headers")
            return

        # Read data rows into columns
        columns: dict[int, list] = {i: [] for i in range(len(headers))}
        row_count = 0

        for row in rows_iter:
            if row_count >= max_rows:
                result.warnings.append(
                    f"Sheet '{sheet_name}' truncated at {max_rows} rows"
                )
                break
            for i, val in enumerate(row):
                if i < len(headers):
                    columns[i].append(val)
            row_count += 1

        # Profile and create elements
        self._profile_columns(headers, columns, row_count, sheet_name, result)

        result.sheets.append({
            "name": sheet_name,
            "row_count": row_count,
            "column_count": len(headers),
            "headers": headers,
        })

    def _profile_columns(
        self,
        headers: list[str],
        columns: dict[int, list],
        row_count: int,
        sheet_name: str,
        result: SpreadsheetParseResult,
    ) -> None:
        """Profile columns and create ontology elements."""
        entity_id = _slug(sheet_name)
        entity_name = _display_name(sheet_name)

        # Create entity for this sheet/table
        entity = {
            "id": entity_id,
            "name": entity_name,
            "description": f"Data table from spreadsheet sheet '{sheet_name}'",
            "core_attributes": [],
            "lenses": [],
            "state": "as-is",
        }

        for i, header in enumerate(headers):
            if not header:
                continue

            col_values = columns.get(i, [])
            non_null = [v for v in col_values if v is not None and str(v).strip() != ""]
            null_count = len(col_values) - len(non_null)

            # Type inference
            data_type = _infer_type_from_values(non_null[:500])

            # Cardinality
            try:
                unique_values = set(str(v) for v in non_null)
                unique_count = len(unique_values)
            except TypeError:
                unique_count = len(non_null)

            # Classification
            classification = _classify_column(
                header, data_type, unique_count, row_count, null_count,
            )

            # Sample values (distinct, up to MAX_SAMPLE_VALUES)
            sample_values = []
            seen = set()
            for v in non_null:
                sv = str(v)
                if sv not in seen and len(sample_values) < MAX_SAMPLE_VALUES:
                    sample_values.append(sv)
                    seen.add(sv)

            # Statistics for numeric columns
            stats = {}
            if data_type == "number":
                numeric_vals = []
                for v in non_null:
                    try:
                        numeric_vals.append(float(str(v).replace(",", "")))
                    except (ValueError, TypeError):
                        pass
                if numeric_vals:
                    stats = {
                        "min": min(numeric_vals),
                        "max": max(numeric_vals),
                        "mean": sum(numeric_vals) / len(numeric_vals),
                    }

            # Column profile (for review UI)
            attr_id = _slug(f"{entity_id}_{header}")
            profile = {
                "column_name": header,
                "sheet_name": sheet_name,
                "data_type": data_type,
                "classification": classification,
                "unique_count": unique_count,
                "null_count": null_count,
                "total_count": row_count,
                "null_percentage": round(null_count / row_count * 100, 1) if row_count > 0 else 0,
                "sample_values": sample_values,
                "statistics": stats,
                "attribute_id": attr_id,
            }
            result.column_profiles.append(profile)

            # Create ontology element based on classification
            if classification == "measure":
                # Numeric measures
                measure = {
                    "id": attr_id,
                    "name": _display_name(header),
                    "description": f"Numeric measure from {sheet_name}.{header}",
                    "logic": f"Aggregation of {header}",
                    "formula": "",
                    "input_attribute_ids": [],
                    "input_measure_ids": [],
                    "perspective_ids": [],
                    "state": "as-is",
                    "source_sheet": sheet_name,
                    "source_column": header,
                    "data_type": data_type,
                    "statistics": stats,
                }
                result.measures.append(measure)
            else:
                # Dimensions and date dimensions → attributes
                attribute = {
                    "id": attr_id,
                    "name": _display_name(header),
                    "description": f"{'Date dimension' if classification == 'date_dimension' else 'Dimension'} from {sheet_name}.{header}",
                    "entity_id": entity_id,
                    "system_id": "",  # Set by _add_system
                    "source_table": sheet_name,
                    "source_column": header,
                    "data_type": data_type,
                    "state": "as-is",
                    "unique_count": unique_count,
                }
                result.attributes.append(attribute)
                entity["core_attributes"].append(attr_id)

        result.entities.append(entity)

    def _add_system(self, result: SpreadsheetParseResult, filename: str) -> None:
        """Add the spreadsheet file as a source system."""
        ext = Path(filename).suffix.lower()
        system_type = "Excel Workbook" if ext in (".xlsx", ".xls") else "CSV File"

        system_id = _slug(f"spreadsheet_{Path(filename).stem}")
        result.systems.append({
            "id": system_id,
            "name": f"{Path(filename).name}",
            "type": "Spreadsheet",
            "vendor": "Microsoft" if ext in (".xlsx", ".xls") else "",
            "integration_status": "Manual Extract",
            "state": "as-is",
            "notes": f"{system_type} used as data source",
        })

        # Backfill system_id on attributes
        for attr in result.attributes:
            if not attr.get("system_id"):
                attr["system_id"] = system_id

    @staticmethod
    def _coerce_csv_value(val: str):
        """Try to coerce a CSV string value to a native Python type."""
        if not val or val.strip() == "":
            return None

        stripped = val.strip()

        # Boolean
        if stripped.lower() in ("true", "false"):
            return stripped.lower() == "true"

        # Integer
        try:
            # Don't coerce things that look like IDs/codes with leading zeros
            if not stripped.startswith("0") or stripped == "0":
                int_val = int(stripped)
                return int_val
        except ValueError:
            pass

        # Float (handle comma as thousands separator)
        try:
            clean = stripped.replace(",", "")
            float_val = float(clean)
            return float_val
        except ValueError:
            pass

        # Date detection (common formats)
        import datetime
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.datetime.strptime(stripped, fmt)
            except ValueError:
                continue

        return stripped
