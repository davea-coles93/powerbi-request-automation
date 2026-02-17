"""Discovery router for Phase 1 data import and workshop sessions."""
import csv
import io
import json
import uuid
from typing import List, Optional, Literal

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.db.database import get_db
from app.db.repositories import (
    PerspectiveRepository,
    SystemRepository,
    EntityRepository,
    AttributeRepository,
    MeasureRepository,
    MetricRepository,
    ProcessRepository,
    WorkshopSessionRepository,
)
from app.models import (
    Perspective,
    System,
    Entity,
    Attribute,
    Measure,
    Metric,
    Process,
    ProcessStep,
)
from app.models.workshop import WorkshopSession, WorkshopFinding


router = APIRouter(prefix="/api/discovery", tags=["discovery"])


# --- Response Models ---

class ImportResult(BaseModel):
    """Result of a CSV import operation."""
    success: bool
    entity_type: str
    items_created: int
    errors: list[str] = Field(default_factory=list)


class ExcelImportResult(BaseModel):
    """Result of an Excel import operation."""
    success: bool
    sheets_processed: list[str] = Field(default_factory=list)
    summary: dict = Field(default_factory=dict)  # sheet_name -> count
    errors: list[str] = Field(default_factory=list)


class WorkshopSessionCreate(BaseModel):
    """Request body for creating a workshop session."""
    name: str
    date: str
    participants: list[str] = Field(default_factory=list)
    session_type: Literal["top_down", "bottom_up", "gap_analysis"] = Field(alias="type")
    notes: Optional[str] = None

    class Config:
        populate_by_name = True


class WorkshopSessionUpdate(BaseModel):
    """Request body for updating a workshop session."""
    name: Optional[str] = None
    date: Optional[str] = None
    participants: Optional[list[str]] = None
    notes: Optional[str] = None


class FindingCreate(BaseModel):
    """Request body for adding a finding to a session."""
    category: Literal[
        "missing_supply",
        "unused_supply",
        "shadow_system",
        "high_manual_effort",
        "data_quality",
        "other",
    ]
    description: str
    related_entity_ids: list[str] = Field(default_factory=list)
    related_attribute_ids: list[str] = Field(default_factory=list)
    priority: Literal["high", "medium", "low"] = "medium"


# --- Helpers ---

# Map entity type names to their repository, Pydantic model, and required fields
ENTITY_TYPE_MAP = {
    "perspectives": {
        "repo_class": PerspectiveRepository,
        "model_class": Perspective,
        "required_fields": ["id", "name", "purpose", "primary_concern"],
    },
    "systems": {
        "repo_class": SystemRepository,
        "model_class": System,
        "required_fields": ["id", "name", "type"],
    },
    "entities": {
        "repo_class": EntityRepository,
        "model_class": Entity,
        "required_fields": ["id", "name"],
    },
    "attributes": {
        "repo_class": AttributeRepository,
        "model_class": Attribute,
        "required_fields": ["id", "name", "entity_id", "system_id"],
    },
    "measures": {
        "repo_class": MeasureRepository,
        "model_class": Measure,
        "required_fields": ["id", "name"],
    },
    "metrics": {
        "repo_class": MetricRepository,
        "model_class": Metric,
        "required_fields": ["id", "name", "business_question"],
    },
    "processes": {
        "repo_class": ProcessRepository,
        "model_class": Process,
        "required_fields": ["id", "name"],
    },
}

# Sheet name normalization: map various sheet names to entity types
SHEET_NAME_MAP = {
    "perspectives": "perspectives",
    "perspective": "perspectives",
    "systems": "systems",
    "system": "systems",
    "entities": "entities",
    "entity": "entities",
    "attributes": "attributes",
    "attribute": "attributes",
    "measures": "measures",
    "measure": "measures",
    "metrics": "metrics",
    "metric": "metrics",
    "processes": "processes",
    "process": "processes",
}


def _generate_id(name: str) -> str:
    """Generate a slug-style ID from a name."""
    return name.lower().strip().replace(" ", "_").replace("-", "_")


def _parse_json_field(value: str) -> Optional[list]:
    """Try to parse a string as JSON list. Returns None if not valid JSON."""
    if not value or not value.strip():
        return None
    value = value.strip()
    # Try JSON parse
    try:
        parsed = json.loads(value)
        if isinstance(parsed, list):
            return parsed
        return [parsed]
    except (json.JSONDecodeError, TypeError):
        pass
    # Try comma-separated
    if "," in value:
        return [v.strip() for v in value.split(",") if v.strip()]
    # Single value
    return [value]


def _row_to_model(row: dict, column_mapping: dict, entity_type: str) -> dict:
    """Convert a CSV row to a model dict using column mapping.

    column_mapping maps CSV column names to model field names.
    e.g., {"System Name": "name", "System Type": "type"}
    """
    model_data = {}
    for csv_col, model_field in column_mapping.items():
        if csv_col in row and row[csv_col] is not None:
            value = str(row[csv_col]).strip()
            if not value:
                continue

            # Handle JSON/list fields
            list_fields = [
                "typical_actors", "consumes_from", "feeds",
                "core_attributes", "lenses", "input_attribute_ids",
                "input_measure_ids", "perspective_ids",
                "calculated_by_measure_ids", "steps",
                "related_entity_ids", "related_attribute_ids",
                "participants", "findings", "constraints",
                "consumes_attribute_ids", "produces_attribute_ids",
                "uses_metric_ids", "crystallizes_attribute_ids",
                "depends_on_step_ids", "systems_used_ids",
                "observations",
            ]
            if model_field in list_fields:
                parsed = _parse_json_field(value)
                if parsed is not None:
                    model_data[model_field] = parsed
            else:
                model_data[model_field] = value

    # Auto-generate ID if not provided
    if "id" not in model_data and "name" in model_data:
        model_data["id"] = _generate_id(model_data["name"])

    return model_data


def _auto_detect_mapping(headers: list[str], entity_type: str) -> dict:
    """Auto-detect column mapping from header names.

    Returns a dict mapping CSV column names to model field names.
    Uses fuzzy matching on common patterns.
    """
    type_info = ENTITY_TYPE_MAP.get(entity_type)
    if not type_info:
        return {}

    model_class = type_info["model_class"]
    model_fields = set(model_class.model_fields.keys())

    mapping = {}
    for header in headers:
        # Exact match
        clean = header.strip().lower().replace(" ", "_").replace("-", "_")
        if clean in model_fields:
            mapping[header] = clean
            continue

        # Common aliases
        aliases = {
            "system_name": "name",
            "entity_name": "name",
            "attribute_name": "name",
            "measure_name": "name",
            "metric_name": "name",
            "process_name": "name",
            "perspective_name": "name",
            "desc": "description",
            "system_type": "type",
            "sys_type": "type",
            "question": "business_question",
            "biz_question": "business_question",
            "source": "source_actor",
            "actor": "source_actor",
            "table": "source_table",
            "column": "source_column",
            "connection": "source_connection",
            "entity": "entity_id",
            "system": "system_id",
            "vendor_name": "vendor",
            "reliability_level": "reliability",
            "integration": "integration_status",
            "formula_text": "formula",
            "logic_description": "logic",
        }
        if clean in aliases and aliases[clean] in model_fields:
            mapping[header] = aliases[clean]

    return mapping


# --- CSV Import ---

@router.post("/import/csv", response_model=ImportResult)
async def import_csv(
    file: UploadFile = File(...),
    entity_type: str = Form(...),
    column_mapping: str = Form(default="{}"),
    db: Session = Depends(get_db),
):
    """Import a CSV file with flexible column mapping.

    Adds data incrementally (does NOT clear existing data).
    """
    if entity_type not in ENTITY_TYPE_MAP:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid entity_type '{entity_type}'. Must be one of: {list(ENTITY_TYPE_MAP.keys())}",
        )

    # Read the CSV content
    try:
        content = await file.read()
        text = content.decode("utf-8-sig")  # Handle BOM in Excel-exported CSVs
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read CSV file: {str(e)}")

    # Parse CSV
    try:
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {str(e)}")

    if not rows:
        return ImportResult(
            success=True,
            entity_type=entity_type,
            items_created=0,
            errors=["CSV file is empty or has no data rows"],
        )

    # Parse column mapping
    try:
        mapping = json.loads(column_mapping)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid column_mapping JSON")

    # Auto-detect mapping if not provided
    if not mapping:
        headers = list(rows[0].keys())
        mapping = _auto_detect_mapping(headers, entity_type)

    if not mapping:
        raise HTTPException(
            status_code=400,
            detail="Could not auto-detect column mapping. Please provide an explicit column_mapping.",
        )

    # Get repository
    type_info = ENTITY_TYPE_MAP[entity_type]
    repo = type_info["repo_class"](db)
    model_class = type_info["model_class"]

    created = 0
    errors = []

    for i, row in enumerate(rows, start=2):  # Start at 2 (header is row 1)
        try:
            model_data = _row_to_model(row, mapping, entity_type)

            # Skip empty rows
            if not model_data or not model_data.get("name"):
                continue

            # Auto-generate ID if missing
            if "id" not in model_data:
                model_data["id"] = _generate_id(model_data["name"])

            # Create the model instance (validates data)
            item = model_class(**model_data)

            # Check for duplicates
            existing = repo.get_by_id(item.id)
            if existing:
                errors.append(f"Row {i}: Item with id '{item.id}' already exists, skipping")
                continue

            repo.create(item)
            created += 1

        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    return ImportResult(
        success=created > 0 or not errors,
        entity_type=entity_type,
        items_created=created,
        errors=errors,
    )


# --- Excel Import ---

@router.post("/import/excel", response_model=ExcelImportResult)
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Import an Excel file with sheet-per-entity-type convention.

    Each sheet name maps to an entity type (e.g., "Systems" -> systems).
    First row is headers, maps to model fields.
    Imports incrementally (adds to existing data, doesn't clear).
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Run: pip install openpyxl",
        )

    # Read the Excel file
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    sheets_processed = []
    summary = {}
    all_errors = []

    for sheet_name in wb.sheetnames:
        # Normalize sheet name to entity type
        normalized = sheet_name.strip().lower().replace(" ", "_").replace("-", "_")
        entity_type = SHEET_NAME_MAP.get(normalized)

        if not entity_type:
            all_errors.append(f"Sheet '{sheet_name}': unrecognized entity type, skipping")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            all_errors.append(f"Sheet '{sheet_name}': no data rows found, skipping")
            continue

        # First row is headers
        headers = [str(h).strip() if h else "" for h in rows[0]]

        # Auto-detect column mapping
        mapping = _auto_detect_mapping(headers, entity_type)
        if not mapping:
            all_errors.append(
                f"Sheet '{sheet_name}': could not detect column mapping for headers {headers}, skipping"
            )
            continue

        type_info = ENTITY_TYPE_MAP[entity_type]
        repo = type_info["repo_class"](db)
        model_class = type_info["model_class"]

        created = 0
        for row_idx, row_data in enumerate(rows[1:], start=2):
            try:
                # Build row dict from headers and values
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    if header and col_idx < len(row_data):
                        value = row_data[col_idx]
                        if value is not None:
                            row_dict[header] = str(value)

                model_data = _row_to_model(row_dict, mapping, entity_type)

                # Skip empty rows
                if not model_data or not model_data.get("name"):
                    continue

                # Auto-generate ID if missing
                if "id" not in model_data:
                    model_data["id"] = _generate_id(model_data["name"])

                item = model_class(**model_data)

                # Check for duplicates
                existing = repo.get_by_id(item.id)
                if existing:
                    all_errors.append(
                        f"Sheet '{sheet_name}', row {row_idx}: id '{item.id}' already exists, skipping"
                    )
                    continue

                repo.create(item)
                created += 1

            except Exception as e:
                all_errors.append(f"Sheet '{sheet_name}', row {row_idx}: {str(e)}")

        sheets_processed.append(sheet_name)
        summary[sheet_name] = created

    wb.close()

    return ExcelImportResult(
        success=len(sheets_processed) > 0,
        sheets_processed=sheets_processed,
        summary=summary,
        errors=all_errors,
    )


# --- Workshop Sessions ---

@router.post("/workshop/sessions", response_model=WorkshopSession)
async def create_workshop_session(
    body: WorkshopSessionCreate,
    db: Session = Depends(get_db),
):
    """Create a new workshop session."""
    repo = WorkshopSessionRepository(db)

    session = WorkshopSession(
        id=str(uuid.uuid4()),
        name=body.name,
        date=body.date,
        participants=body.participants,
        session_type=body.session_type,
        notes=body.notes,
        findings=[],
    )

    return repo.create(session)


@router.get("/workshop/sessions", response_model=List[WorkshopSession])
async def list_workshop_sessions(db: Session = Depends(get_db)):
    """List all workshop sessions."""
    repo = WorkshopSessionRepository(db)
    return repo.get_all()


@router.get("/workshop/sessions/{session_id}", response_model=WorkshopSession)
async def get_workshop_session(session_id: str, db: Session = Depends(get_db)):
    """Get a specific workshop session by ID."""
    repo = WorkshopSessionRepository(db)
    session = repo.get_by_id(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Workshop session not found")
    return session


@router.put("/workshop/sessions/{session_id}", response_model=WorkshopSession)
async def update_workshop_session(
    session_id: str,
    body: WorkshopSessionUpdate,
    db: Session = Depends(get_db),
):
    """Update a workshop session (add notes, update details)."""
    repo = WorkshopSessionRepository(db)
    existing = repo.get_by_id(session_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Workshop session not found")

    # Merge updates
    update_data = existing.model_dump()
    for key, value in body.model_dump(exclude_unset=True).items():
        if value is not None:
            update_data[key] = value

    updated = WorkshopSession(**update_data)
    return repo.update(session_id, updated)


@router.delete("/workshop/sessions/{session_id}")
async def delete_workshop_session(session_id: str, db: Session = Depends(get_db)):
    """Delete a workshop session."""
    repo = WorkshopSessionRepository(db)
    success = repo.delete(session_id)
    if not success:
        raise HTTPException(status_code=404, detail="Workshop session not found")
    return {"message": f"Workshop session {session_id} deleted successfully"}


@router.post("/workshop/sessions/{session_id}/findings", response_model=WorkshopSession)
async def add_finding(
    session_id: str,
    body: FindingCreate,
    db: Session = Depends(get_db),
):
    """Add a finding to a workshop session."""
    repo = WorkshopSessionRepository(db)
    session = repo.get_by_id(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Workshop session not found")

    # Create the finding
    finding = WorkshopFinding(
        id=str(uuid.uuid4()),
        category=body.category,
        description=body.description,
        related_entity_ids=body.related_entity_ids,
        related_attribute_ids=body.related_attribute_ids,
        priority=body.priority,
    )

    # Add to session findings
    session.findings.append(finding)

    # Update the session
    return repo.update(session_id, session)


@router.delete(
    "/workshop/sessions/{session_id}/findings/{finding_id}",
    response_model=WorkshopSession,
)
async def remove_finding(
    session_id: str,
    finding_id: str,
    db: Session = Depends(get_db),
):
    """Remove a finding from a workshop session."""
    repo = WorkshopSessionRepository(db)
    session = repo.get_by_id(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Workshop session not found")

    original_count = len(session.findings)
    session.findings = [f for f in session.findings if f.id != finding_id]

    if len(session.findings) == original_count:
        raise HTTPException(status_code=404, detail="Finding not found")

    return repo.update(session_id, session)
